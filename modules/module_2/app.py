"""
Created Histories Module - Flet-based Modern UI
Device History Management with dashboard, advanced filtering, and professional design
"""

import flet as ft
import csv
import hashlib
import logging
import os
import sys
from datetime import datetime
from typing import Dict, List, Any, Optional

# Add parent directory to path for imports
sys.path.insert(0, os. path.dirname(os.path. abspath(__file__)))

# Optional: openpyxl for Excel import
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from inventory_db import InventoryDatabase
from ui_components. theme_flet import ThemeManager, get_status_colors
from ui_components.dashboard_cards_flet import MetricCard, StatisticsCard
from ui_components. enhanced_grid_flet import EnhancedDataGrid
from ui_components.status_indicators_flet import StatusIndicator
from ui_components. oor_parser import OORParser

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class CreatedHistoriesApp:
    """Modern Created Histories Application using Flet"""
    
    # Master password hash (default: "admin123")
    MASTER_PASSWORD_HASH = "0192023a7bbd73250516f069df18b500"
    
    SHEETS = [
        ("Ohio", "Meters"),
        ("I&M", "Meters"),
        ("Ohio", "Transformers"),
        ("I&M", "Transformers"),
    ]
    
    # Full database columns
    FULL_COLUMNS = [
        "ID", "OpCo", "Device Type", "Status", "MFR", "Dev Code", "Beg Ser", "End Ser",
        "OOR Serial", "Qty", "PO Date", "PO Number", "Recv Date", "Unit Cost", "CID",
        "M. E.  #", "Pur. Code", "Est.", "Use", "Notes 1", "Notes 2"
    ]
    
    # Streamlined grid columns for display
    GRID_COLUMNS = [
        "ID", "Dev Code", "Beg Ser", "End Ser", "OOR Serial", "Qty", "Recv Date", "CID"
    ]
    
    # Column index mapping from full to grid columns
    COLUMN_INDEX_MAP = {
        "ID": 0,
        "Dev Code": 5,
        "Beg Ser": 6,
        "End Ser": 7,
        "OOR Serial": 8,
        "Qty": 9,
        "Recv Date": 12,
        "CID": 14,
    }
    
    # Sheet name mapping for import
    SHEET_MAP = {
        'OH - Meters': ('Ohio', 'Meters'),
        'I&M - Meters': ('I&M', 'Meters'),
        'OH - Transformers': ('Ohio', 'Transformers'),
        'I&M - Transformers': ('I&M', 'Transformers'),
    }
    
    def __init__(self, page: ft.Page):
        self. page = page
        self.edit_mode = False
        self.current_theme = "Dark"
        self. current_sheet: Optional[str] = None
        self. current_view = "dashboard"
        self. selected_record_data: Optional[Dict] = None
        
        # Initialize database
        self.db = InventoryDatabase('created_histories.db')
        self.db.init_db()
        
        # UI component references
        self.nav_rail: Optional[ft.NavigationRail] = None
        self. content_area: Optional[ft. Container] = None
        self.detail_panel: Optional[ft. Container] = None
        self.breadcrumb_label: Optional[ft. Text] = None
        self.subtitle_label: Optional[ft.Text] = None
        self.mode_indicator: Optional[ft.Container] = None
        self.edit_mode_btn: Optional[ft. TextButton] = None
        self.search_field: Optional[ft. TextField] = None
        
        # Dashboard cards
        self. total_records_card: Optional[MetricCard] = None
        self.total_devices_card: Optional[MetricCard] = None
        self. total_value_card: Optional[MetricCard] = None
        self. avg_cost_card: Optional[MetricCard] = None
        self. sheet_stats_cards: Dict[str, StatisticsCard] = {}
        
        # Sheet views
        self.sheet_grids: Dict[str, EnhancedDataGrid] = {}
        
        # Setup page
        self.setup_page()
        
        # Build UI
        self. build_ui()
        
        logger.info("Created Histories App initialized!")
    
    def setup_page(self):
        """Configure page settings"""
        self. page.title = "Created Histories - Device History Management"
        self.page.window. width = 1600
        self.page.window.height = 1075
        self.page.padding = 0
        self.page.theme_mode = ft.ThemeMode.DARK
        self.page.theme = ft.Theme(color_scheme_seed=ft.Colors.BLUE)
    
    def build_ui(self):
        """Build the main user interface"""
        app_bar = self.create_app_bar()
        self.nav_rail = self. create_navigation_rail()
        
        self.content_area = ft.Container(
            content=self.create_dashboard_view(),
            expand=True,
            padding=24,
        )
        
        main_content = ft.Row(
            controls=[
                self. nav_rail,
                ft. VerticalDivider(width=1),
                self.content_area,
            ],
            expand=True,
            spacing=0,
        )
        
        self. page.appbar = app_bar
        self.page.add(main_content)

    def create_app_bar(self) -> ft.AppBar:
        """Create the application bar"""
        self.breadcrumb_label = ft.Text(
            "Dashboard",
            size=18,
            weight=ft.FontWeight.BOLD,
        )
        self.subtitle_label = ft.Text(
            "Overview of all device histories",
            size=12,
            color=ft.Colors.ON_SURFACE_VARIANT,
        )

        breadcrumb_column = ft.Column(
            controls=[self.breadcrumb_label, self.subtitle_label],
            spacing=0,
        )

        self.mode_indicator = StatusIndicator('inactive', 'Read-Only')

        self.edit_mode_btn = ft.TextButton(
            text="🔓 Enable Edit Mode",
            on_click=self.toggle_edit_mode,
        )

        return ft.AppBar(
            leading=ft.Icon(ft.Icons.HISTORY_EDU),
            leading_width=50,
            title=breadcrumb_column,
            center_title=False,
            bgcolor=ft.Colors.SURFACE_CONTAINER_HIGHEST,
            actions=[
                self.mode_indicator,
                ft.Container(width=16),
                self.edit_mode_btn,
                ft.Container(width=16),
                ft.IconButton(
                    icon=ft.Icons.DARK_MODE,
                    tooltip="Toggle Theme",
                    on_click=self.toggle_theme,
                ),
                ft.PopupMenuButton(
                    icon=ft.Icons.MORE_VERT,
                    items=[
                        ft.PopupMenuItem(
                            text="About",
                            icon=ft.Icons.INFO_OUTLINE,
                            on_click=self.show_about,
                        ),
                        ft.PopupMenuItem(
                            text="Help",
                            icon=ft.Icons.HELP_OUTLINE,
                            on_click=self.show_help,
                        ),
                    ],
                ),
            ],
        )

    def create_navigation_rail(self) -> ft.NavigationRail:
        """Create the navigation rail"""
        destinations = [
            ft.NavigationRailDestination(
                icon=ft.Icons.DASHBOARD_OUTLINED,
                selected_icon=ft.Icons.DASHBOARD,
                label="Dashboard",
            ),
        ]

        sheet_labels = ["Ohio Meters", "I&M Meters", "Ohio Transformers", "I&M Transformers"]

        for idx, (opco, device_type) in enumerate(self.SHEETS):
            destinations.append(
                ft.NavigationRailDestination(
                    icon=ft.Icons.TABLE_CHART_OUTLINED,
                    selected_icon=ft.Icons.TABLE_CHART,
                    label=sheet_labels[idx],
                )
            )

        return ft.NavigationRail(
            selected_index=0,
            label_type=ft.NavigationRailLabelType.ALL,
            min_width=80,  # Reduced from 100
            min_extended_width=180,
            destinations=destinations,
            on_change=self.on_nav_change,
        )
    
    def on_nav_change(self, e):
        """Handle navigation rail selection changes"""
        index = e.control. selected_index
        logger.info(f"Navigation changed to index: {index}")
        
        if index == 0:
            self.show_dashboard()
        else:
            sheet_idx = index - 1
            if 0 <= sheet_idx < len(self. SHEETS):
                opco, device_type = self.SHEETS[sheet_idx]
                sheet_name = f"{opco} - {device_type}"
                self.show_sheet(sheet_name)
    
    def on_search_change(self, e):
        """Handle search field changes"""
        search_text = e. control.value. strip(). lower()
        
        if self.current_view == "sheet" and self.current_sheet:
            opco, device_type = None, None
            for o, d in self. SHEETS:
                if f"{o} - {d}" == self.current_sheet:
                    opco, device_type = o, d
                    break
            
            if opco and device_type:
                self.search_sheet_data(opco, device_type, search_text)
    
    def search_sheet_data(self, opco: str, device_type: str, search_text: str):
        """Search data across all columns"""
        sheet_name = f"{opco} - {device_type}"
        grid = self.sheet_grids.get(sheet_name)
        
        if not grid:
            return
        
        items = self.db.get_items_by_sheet(opco, device_type)
        
        if search_text:
            filtered_items = []
            for item in items:
                if any(search_text in str(val).lower() for val in item if val):
                    grid_item = self._convert_to_grid_format(item)
                    filtered_items. append(grid_item)
            grid.set_data(filtered_items)
        else:
            grid_data = [self._convert_to_grid_format(item) for item in items]
            grid.set_data(grid_data)
    
    def _convert_to_grid_format(self, full_record: tuple) -> tuple:
        """Convert full database record to streamlined grid format"""
        grid_record = []
        for col_name in self.GRID_COLUMNS:
            idx = self.COLUMN_INDEX_MAP. get(col_name, 0)
            grid_record. append(full_record[idx] if idx < len(full_record) else "")
        return tuple(grid_record)
    
    def create_dashboard_view(self) -> ft. Control:
        """Create the dashboard view with metrics"""
        self.total_records_card = MetricCard("Total Records", "0", "📊")
        self.total_devices_card = MetricCard("Total Devices", "0", "🔧")
        self. total_value_card = MetricCard("Total Value", "$0", "💰")
        self. avg_cost_card = MetricCard("Avg.  Unit Cost", "$0", "📈")
        
        metrics_row = ft.Row(
            controls=[
                self. total_records_card,
                self. total_devices_card,
                self. total_value_card,
                self. avg_cost_card,
            ],
            spacing=16,
            wrap=True,
            run_spacing=16,
        )
        
        sheet_stats_label = ft.Text("Sheet Statistics", size=18, weight=ft.FontWeight.W_600)
        quick_actions_label = ft.Text("Quick Access", size=18, weight=ft. FontWeight.W_600)
        
        sheet_labels = [
            ("Ohio M", "Ohio", "Meters"),
            ("I&M M", "I&M", "Meters"),
            ("Ohio T", "Ohio", "Transformers"),
            ("I&M T", "I&M", "Transformers"),
        ]
        
        quick_action_buttons = ft.Row(
            controls=[
                ft.ElevatedButton(
                    text=label,
                    icon=ft.Icons. TABLE_CHART,
                    on_click=lambda e, o=opco, d=dtype: self.nav_to_sheet(o, d),
                    width=150,
                )
                for label, opco, dtype in sheet_labels
            ],
            spacing=16,
            wrap=True,
        )
        
        sheet_stats_row = ft.Row(controls=[], spacing=16, wrap=True, run_spacing=16)
        
        for opco, device_type in self.SHEETS:
            sheet_name = f"{opco} - {device_type}"
            card = StatisticsCard(sheet_name, {})
            self.sheet_stats_cards[sheet_name] = card
            sheet_stats_row. controls.append(card)
        
        dashboard = ft.Column(
            controls=[
                ft.Text("Dashboard", size=24, weight=ft.FontWeight. BOLD),
                ft.Container(height=16),
                metrics_row,
                ft.Container(height=24),
                quick_actions_label,
                ft.Container(height=12),
                quick_action_buttons,
                ft.Container(height=24),
                sheet_stats_label,
                ft. Container(height=16),
                sheet_stats_row,
            ],
            scroll=ft.ScrollMode.AUTO,
            expand=True,
        )
        
        self.update_dashboard_metrics()
        return dashboard
    
    def nav_to_sheet(self, opco: str, device_type: str):
        """Navigate to a sheet from dashboard button"""
        sheet_name = f"{opco} - {device_type}"
        for idx, (o, d) in enumerate(self.SHEETS):
            if o == opco and d == device_type:
                self.nav_rail.selected_index = idx + 1
                self.nav_rail. update()
                break
        self.show_sheet(sheet_name)

    def create_sheet_view(self, opco: str, device_type: str) -> ft.Control:
        """Create a sheet view with data grid, toolbar, and detail panel"""
        sheet_name = f"{opco} - {device_type}"

        # Create data grid
        grid = EnhancedDataGrid(
            columns=self.GRID_COLUMNS,
            on_row_select=lambda idx, o=opco, d=device_type: self.on_row_selected(o, d, idx),
            on_row_double_click=lambda idx, o=opco, d=device_type: self.edit_record(o, d),
        )
        self.sheet_grids[sheet_name] = grid

        # Search field
        search_field = ft.TextField(
            hint_text="Search across all columns.. .",
            expand=True,
            height=40,
            dense=True,
            border_radius=20,
            prefix_icon=ft.Icons.SEARCH,
            on_change=self.on_search_change,
            bgcolor=ft.Colors.SURFACE_CONTAINER_HIGHEST,
        )

        # Toolbar with search in the middle
        toolbar = ft.Row(
            controls=[
                ft.FilledButton(
                    text="Add Record",
                    icon=ft.Icons.ADD,
                    on_click=lambda e: self.add_record(opco, device_type),
                ),
                ft.OutlinedButton(
                    text="Delete",
                    icon=ft.Icons.DELETE,
                    style=ft.ButtonStyle(color=ft.Colors.ERROR),
                    on_click=lambda e: self.delete_record(opco, device_type),
                ),
                ft.TextField(
                    hint_text="Search.. .",
                    width=300,
                    height=40,
                    dense=True,
                    border_radius=20,
                    prefix_icon=ft.Icons.SEARCH,
                    on_change=self.on_search_change,
                    bgcolor=ft.Colors.SURFACE_CONTAINER_HIGHEST,
                ),
                ft.Container(expand=True),
                ft.OutlinedButton(
                    text="Import",
                    icon=ft.Icons.FILE_UPLOAD,
                    on_click=lambda e: self.import_data(opco, device_type),
                ),
                ft.OutlinedButton(
                    text="Export",
                    icon=ft.Icons.FILE_DOWNLOAD,
                    on_click=lambda e: self.export_data(opco, device_type),
                ),
                ft.OutlinedButton(
                    text="Statistics",
                    icon=ft.Icons.BAR_CHART,
                    on_click=lambda e: self.show_statistics(opco, device_type),
                ),
            ],
            spacing=8,
        )

        # Create detail panel
        self.detail_panel = self.create_detail_panel()

        # Left section with toolbar and grid
        left_section = ft.Column(
            controls=[
                toolbar,
                ft.Container(height=8),
                grid,
            ],
            spacing=0,
            expand=True,
        )

        # Main layout
        main_content = ft.Row(
            controls=[
                left_section,
                self.detail_panel,
            ],
            spacing=0,
            expand=True,
        )

        # Load data
        self.load_sheet_data(opco, device_type)

        return main_content

    def create_detail_panel(self) -> ft.Container:
        """Create the detail panel for showing selected record details"""
        return ft.Container(
            content=ft.Column(
                controls=[
                    ft.Row([
                        ft.Text("Record Details", size=18, weight=ft.FontWeight.BOLD),
                        ft.Container(expand=True),
                    ]),
                    ft.Divider(),
                    ft.Container(height=40),
                    ft.Icon(ft.Icons.TOUCH_APP, size=64, color=ft.Colors.ON_SURFACE_VARIANT),
                    ft.Container(height=16),
                    ft.Text(
                        "Select a record to view details",
                        size=14,
                        color=ft.Colors.ON_SURFACE_VARIANT,
                        italic=True,
                        text_align=ft.TextAlign.CENTER,
                    ),
                ],
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                scroll=ft.ScrollMode.AUTO,
                spacing=4,
            ),
            width=480,
            expand=True,  # This makes it stretch vertically
            padding=16,
            bgcolor=ft.Colors.SURFACE_CONTAINER_HIGHEST,
            border=ft.border.only(left=ft.BorderSide(1, ft.Colors.OUTLINE_VARIANT)),
        )

    def update_detail_panel(self, record_data):
        """Update detail panel with selected record data - matching Edit dialog layout"""
        if not self.detail_panel:
            return

        if not record_data:
            # Reset to empty state
            self.detail_panel.content = ft.Column(
                controls=[
                    ft.Row([
                        ft.Text("Record Details", size=18, weight=ft.FontWeight.BOLD),
                        ft.Container(expand=True),
                    ]),
                    ft.Divider(),
                    ft.Container(height=40),
                    ft.Icon(ft.Icons.TOUCH_APP, size=64, color=ft.Colors.ON_SURFACE_VARIANT),
                    ft.Container(height=16),
                    ft.Text(
                        "Select a record to view details",
                        size=14,
                        color=ft.Colors.ON_SURFACE_VARIANT,
                        italic=True,
                        text_align=ft.TextAlign.CENTER,
                    ),
                ],
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                scroll=ft.ScrollMode.AUTO,
                spacing=4,
            )
            if self.detail_panel.page:
                self.detail_panel.update()
            return

        self.selected_record_data = record_data
        is_editable = self.edit_mode

        # Helper to get value safely
        def get_val(key, default=''):
            val = record_data.get(key, default) if isinstance(record_data, dict) else default
            return str(val) if val is not None else ''

        # Format unit cost
        unit_cost = record_data.get('unit_cost', 0) if isinstance(record_data, dict) else 0
        unit_cost_str = f"{float(unit_cost):.2f}" if unit_cost else "0. 00"

        # Format qty
        qty = record_data.get('qty', 0) if isinstance(record_data, dict) else 0
        qty_str = f"{int(qty):,}" if qty else "0"

        record_id = get_val('id', 'N/A')

        # Build form layout matching the Edit dialog exactly
        form_content = ft.Column(
            controls=[
                # Header with ID badge
                ft.Row([
                    ft.Text("Record Details", size=18, weight=ft.FontWeight.BOLD),
                    ft.Container(expand=True),
                    ft.Container(
                        content=ft.Text(f"ID: {record_id}", size=12, weight=ft.FontWeight.W_500),
                        padding=ft.padding.symmetric(horizontal=12, vertical=6),
                        border_radius=16,
                        bgcolor=ft.Colors.PRIMARY_CONTAINER,
                    ),
                ]),
                ft.Divider(),
                ft.Container(height=4),

                # Row 1: OpCo and Device Type (like Edit dialog)
                ft.Row([
                    ft.TextField(label="OpCo", value=get_val('opco'), read_only=True, disabled=True, expand=True),
                    ft.TextField(label="Device Type", value=get_val('device_type'), read_only=True, disabled=True,
                                 expand=True),
                ], spacing=12),

                # Row 2: Status and Manufacturer (like Edit dialog)
                ft.Row([
                    ft.TextField(label="Status", value=get_val('status'), read_only=not is_editable, expand=True),
                    ft.TextField(label="Manufacturer", value=get_val('mfr'), read_only=not is_editable, expand=True),
                ], spacing=12),

                # Row 3: Device Code, Beginning Serial, Ending Serial (3 columns like Edit dialog)
                ft.Row([
                    ft.TextField(label="Device Code", value=get_val('dev_code'), read_only=not is_editable,
                                 expand=True),
                    ft.TextField(label="Beginning Serial", value=get_val('beg_ser'), read_only=not is_editable,
                                 expand=True),
                    ft.TextField(label="Ending Serial", value=get_val('end_ser'), read_only=not is_editable,
                                 expand=True),
                ], spacing=12),

                # OOR Serial (full width like Edit dialog)
                ft.TextField(
                    label="OOR Serial",
                    value=get_val('oor_serial') or "",
                    read_only=not is_editable,
                    multiline=True,
                    min_lines=1,
                    max_lines=2,
                ),

                # Note about Qty auto-calculate (like Edit dialog)
                ft.Text(
                    "Note: Qty will auto-calculate from OOR Serial or Beg/End Ser",
                    size=11,
                    italic=True,
                    color=ft.Colors.ON_SURFACE_VARIANT,
                ),

                # Row 4: Quantity, PO Date, PO Number (3 columns like Edit dialog)
                ft.Row([
                    ft.TextField(label="Quantity", value=qty_str, read_only=True, expand=True),
                    ft.TextField(label="PO Date", value=get_val('po_date'), read_only=not is_editable, expand=True),
                    ft.TextField(label="PO Number", value=get_val('po_number'), read_only=not is_editable, expand=True),
                ], spacing=12),

                # Row 5: Received Date, Unit Cost, CID (3 columns like Edit dialog)
                ft.Row([
                    ft.TextField(label="Received Date", value=get_val('recv_date'), read_only=not is_editable,
                                 expand=True),
                    ft.TextField(label="Unit Cost", value=unit_cost_str, read_only=not is_editable, expand=True),
                    ft.TextField(label="CID", value=get_val('cid'), read_only=not is_editable, expand=True),
                ], spacing=12),

                # Row 6: M. E. #, Purchase Code, Est., Use (4 columns like Edit dialog)
                ft.Row([
                    ft.TextField(label="M.E. #", value=get_val('me_number'), read_only=not is_editable, expand=True),
                    ft.TextField(label="Purchase Code", value=get_val('pur_code'), read_only=not is_editable,
                                 expand=True),
                    ft.TextField(label="Est.", value=get_val('est'), read_only=not is_editable, expand=True),
                    ft.TextField(label="Use", value=get_val('use'), read_only=not is_editable, expand=True),
                ], spacing=12),

                # Notes 1 (full width like Edit dialog)
                ft.TextField(
                    label="Notes 1",
                    value=get_val('notes1'),
                    read_only=not is_editable,
                    multiline=True,
                    min_lines=2,
                    max_lines=3,
                ),

                # Notes 2 (full width like Edit dialog)
                ft.TextField(
                    label="Notes 2",
                    value=get_val('notes2'),
                    read_only=not is_editable,
                    multiline=True,
                    min_lines=2,
                    max_lines=3,
                ),
            ],
            scroll=ft.ScrollMode.AUTO,
            spacing=12,
        )

        # Add action buttons if in edit mode
        if is_editable:
            form_content.controls.extend([
                ft.Container(height=8),
                ft.Divider(),
                ft.Container(height=8),
                ft.Row([
                    ft.Container(expand=True),
                    ft.OutlinedButton(text="Duplicate", icon=ft.Icons.COPY, on_click=lambda e: self.duplicate_record()),
                    ft.OutlinedButton(text="Delete", icon=ft.Icons.DELETE, style=ft.ButtonStyle(color=ft.Colors.ERROR),
                                      on_click=lambda e: self.delete_record_from_detail()),
                    ft.FilledButton(text="Edit", icon=ft.Icons.EDIT, on_click=lambda e: self.edit_record_from_detail()),
                ], spacing=8),
            ])

        self.detail_panel.content = form_content
        if self.detail_panel.page:
            self.detail_panel.update()
    
    def on_row_selected(self, opco: str, device_type: str, row_idx: int):
        """Handle row selection to update detail panel"""
        sheet_name = f"{opco} - {device_type}"
        grid = self.sheet_grids. get(sheet_name)
        
        if grid:
            selected_data = grid.get_selected_data()
            if selected_data:
                item_id = selected_data[0]
                full_record = self. db.get_item_by_id(item_id)
                if full_record:
                    self.selected_record_data = full_record
                    self.update_detail_panel(full_record)
            else:
                self.selected_record_data = None
                self.update_detail_panel(None)
    
    def show_dashboard(self):
        """Show dashboard view"""
        self. current_view = "dashboard"
        self.current_sheet = None
        self.selected_record_data = None
        
        if self.breadcrumb_label:
            self. breadcrumb_label.value = "Dashboard"
        if self.subtitle_label:
            self.subtitle_label. value = "Overview of all device histories"
        
        if self.content_area:
            self. content_area.content = self.create_dashboard_view()
        
        self.page.update()
    
    def show_sheet(self, sheet_name: str):
        """Show specific sheet"""
        self.current_view = "sheet"
        self. current_sheet = sheet_name
        self.selected_record_data = None
        
        opco, device_type = None, None
        for o, d in self. SHEETS:
            if f"{o} - {d}" == sheet_name:
                opco, device_type = o, d
                break
        
        if not opco or not device_type:
            logger.error(f"Sheet not found: {sheet_name}")
            return
        
        if self.breadcrumb_label:
            self.breadcrumb_label.value = sheet_name
        if self.subtitle_label:
            self. subtitle_label.value = f"Manage {sheet_name} device records"
        
        if self.content_area:
            self.content_area.content = self.create_sheet_view(opco, device_type)
        
        self.page. update()
    
    def load_sheet_data(self, opco: str, device_type: str):
        """Load data for sheet"""
        try:
            sheet_name = f"{opco} - {device_type}"
            grid = self. sheet_grids. get(sheet_name)
            if not grid:
                return
            
            items = self.db. get_items_by_sheet(opco, device_type)
            grid_data = [self._convert_to_grid_format(item) for item in items]
            grid.set_data(grid_data)
            logger.info(f"Loaded {len(items)} items for {sheet_name}")
            
        except Exception as e:
            logger.error(f"Error loading data: {e}")
            self.show_snackbar(f"Error loading data: {e}", is_error=True)
    
    def update_dashboard_metrics(self):
        """Update dashboard metrics"""
        try:
            total_records = 0
            total_qty = 0
            total_value = 0.0
            costs = []
            
            for opco, device_type in self.SHEETS:
                stats = self.db. get_statistics(opco, device_type)
                
                total_records += stats. get('total_items', 0)
                total_qty += stats.get('total_qty', 0)
                total_value += stats.get('total_value', 0.0)
                
                avg_cost = stats. get('avg_cost', 0.0)
                if avg_cost > 0:
                    costs.append(avg_cost)
                
                sheet_name = f"{opco} - {device_type}"
                card = self.sheet_stats_cards. get(sheet_name)
                if card:
                    card.update_stats({
                        'Records': stats.get('total_items', 0),
                        'Devices': stats.get('total_qty', 0),
                        'Value': f"${stats.get('total_value', 0.0):,.2f}"
                    })
            
            if self.total_records_card:
                self.total_records_card.update_value(f"{total_records:,}")
            if self.total_devices_card:
                self.total_devices_card.update_value(f"{total_qty:,}")
            if self. total_value_card:
                self. total_value_card.update_value(f"${total_value:,.2f}")
            
            avg_cost = sum(costs) / len(costs) if costs else 0
            if self.avg_cost_card:
                self.avg_cost_card.update_value(f"${avg_cost:,.2f}")
            
        except Exception as e:
            logger.error(f"Error updating metrics: {e}")
    
    def toggle_edit_mode(self, e):
        """Toggle edit mode with password"""
        if self. edit_mode:
            self.edit_mode = False
            self.update_mode_indicator()
            if self.edit_mode_btn:
                self. edit_mode_btn.text = "🔓 Enable Edit Mode"
            self.show_snackbar("Edit mode disabled")
            if self.selected_record_data:
                self.update_detail_panel(self.selected_record_data)
        else:
            self.show_password_dialog()
        
        self.page.update()
    
    def update_mode_indicator(self):
        """Update the mode indicator display"""
        if self.mode_indicator:
            if self.edit_mode:
                self. mode_indicator.content = ft.Row(
                    controls=[
                        ft.Icon(ft.Icons. EDIT, size=16, color=ft. Colors.GREEN),
                        ft.Text("Edit Mode", size=12, color=ft.Colors. GREEN),
                    ],
                    spacing=4,
                )
                self.mode_indicator.bgcolor = ft.Colors. GREEN_900
            else:
                self.mode_indicator.content = ft.Row(
                    controls=[
                        ft.Icon(ft.Icons. LOCK_OUTLINE, size=16, color=ft.Colors.ON_SURFACE_VARIANT),
                        ft.Text("Read-Only", size=12, color=ft.Colors.ON_SURFACE_VARIANT),
                    ],
                    spacing=4,
                )
                self.mode_indicator.bgcolor = ft. Colors.SURFACE_CONTAINER_HIGHEST
    
    def show_password_dialog(self):
        """Show password dialog for edit mode"""
        password_field = ft.TextField(
            label="Master Password",
            password=True,
            can_reveal_password=True,
            autofocus=True,
        )
        
        def close_dialog(e):
            dialog. open = False
            self.page.update()
        
        def submit_password(e):
            password = password_field.value or ""
            password_hash = hashlib.md5(password.encode()).hexdigest()
            
            if password_hash == self. MASTER_PASSWORD_HASH:
                self.edit_mode = True
                self.update_mode_indicator()
                if self.edit_mode_btn:
                    self.edit_mode_btn.text = "🔒 Disable Edit Mode"
                self. show_snackbar("Edit mode enabled!")
                dialog.open = False
                if self.selected_record_data:
                    self.update_detail_panel(self.selected_record_data)
            else:
                self. show_snackbar("Incorrect password", is_error=True)
            
            self.page.update()
        
        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text("Enter Password"),
            content=ft.Container(content=password_field, width=300),
            actions=[
                ft.TextButton("Cancel", on_click=close_dialog),
                ft.FilledButton("OK", on_click=submit_password),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        self.page.overlay. append(dialog)
        dialog.open = True
        self.page.update()
    
    def edit_record_from_detail(self):
        """Edit record from detail panel"""
        if self.current_sheet and self.selected_record_data:
            opco, device_type = None, None
            for o, d in self.SHEETS:
                if f"{o} - {d}" == self.current_sheet:
                    opco, device_type = o, d
                    break
            if opco and device_type:
                self.show_add_edit_dialog(opco, device_type, self.selected_record_data)
    
    def duplicate_record(self):
        """Duplicate the selected record"""
        if self.current_sheet and self. selected_record_data:
            opco, device_type = None, None
            for o, d in self.SHEETS:
                if f"{o} - {d}" == self.current_sheet:
                    opco, device_type = o, d
                    break
            if opco and device_type:
                self.show_add_edit_dialog(opco, device_type, self.selected_record_data, is_duplicate=True)
    
    def delete_record_from_detail(self):
        """Delete record from detail panel"""
        if self.current_sheet and self. selected_record_data:
            opco, device_type = None, None
            for o, d in self.SHEETS:
                if f"{o} - {d}" == self.current_sheet:
                    opco, device_type = o, d
                    break
            if opco and device_type:
                self.delete_record(opco, device_type)
    
    def add_record(self, opco: str, device_type: str):
        """Add new record"""
        if not self.edit_mode:
            self.show_snackbar("Edit mode is locked. Enable edit mode first.", is_error=True)
            return
        self.show_add_edit_dialog(opco, device_type, None)
    
    def edit_record(self, opco: str, device_type: str):
        """Edit selected record"""
        if not self.edit_mode:
            self.show_snackbar("Edit mode is locked. Enable edit mode first.", is_error=True)
            return
        
        sheet_name = f"{opco} - {device_type}"
        grid = self. sheet_grids. get(sheet_name)
        if not grid:
            return
        
        row_data = grid.get_selected_data()
        if not row_data:
            self.show_snackbar("Please select a record to edit", is_error=True)
            return
        
        item_id = row_data[0]
        item = self.db. get_item_by_id(item_id)
        
        if not item:
            self.show_snackbar("Could not load record", is_error=True)
            return
        
        self.show_add_edit_dialog(opco, device_type, item)
    
    def show_add_edit_dialog(self, opco: str, device_type: str, item: Optional[Dict] = None, is_duplicate: bool = False):
        """Show add/edit record dialog"""
        is_edit = item is not None and not is_duplicate
        
        status_field = ft. TextField(label="Status", value=item.get('status', '') if item else '', expand=True)
        mfr_field = ft.TextField(label="Manufacturer", value=item.get('mfr', '') if item else '', expand=True)
        dev_code_field = ft.TextField(label="Device Code", value=item. get('dev_code', '') if item else '', expand=True)
        beg_ser_field = ft.TextField(label="Beginning Serial", value=item.get('beg_ser', '') if item else '', expand=True)
        end_ser_field = ft.TextField(label="Ending Serial", value=item.get('end_ser', '') if item else '', expand=True)
        oor_serial_field = ft. TextField(
            label="OOR Serial",
            value=item.get('oor_serial', '') if item else '',
            expand=True,
            multiline=True,
            hint_text="e.g., 1000-1010, 1050, 2000-2005"
        )
        qty_field = ft. TextField(label="Quantity", value=str(item.get('qty', 0)) if item else '0', expand=True)
        po_date_field = ft.TextField(label="PO Date", value=item.get('po_date', '') if item else '', expand=True)
        po_number_field = ft.TextField(label="PO Number", value=item.get('po_number', '') if item else '', expand=True)
        recv_date_field = ft.TextField(label="Received Date", value=item.get('recv_date', '') if item else '', expand=True)
        unit_cost_field = ft.TextField(label="Unit Cost", value=str(item.get('unit_cost', 0.0)) if item else '0.00', expand=True)
        cid_field = ft.TextField(label="CID", value=item. get('cid', '') if item else '', expand=True)
        me_number_field = ft.TextField(label="M.E.  #", value=item.get('me_number', '') if item else '', expand=True)
        pur_code_field = ft.TextField(label="Purchase Code", value=item. get('pur_code', '') if item else '', expand=True)
        est_field = ft.TextField(label="Est.", value=item.get('est', '') if item else '', expand=True)
        use_field = ft.TextField(label="Use", value=item.get('use', '') if item else '', expand=True)
        notes1_field = ft.TextField(label="Notes 1", value=item.get('notes1', '') if item else '', multiline=True, min_lines=2, max_lines=3)
        notes2_field = ft.TextField(label="Notes 2", value=item.get('notes2', '') if item else '', multiline=True, min_lines=2, max_lines=3)
        
        def close_dialog(e):
            dialog.open = False
            self.page.update()
        
        def save_record(e):
            oor_serial_value = oor_serial_field.value or ''
            qty = 0
            
            try:
                if oor_serial_value. strip():
                    parser = OORParser()
                    if parser.parse(oor_serial_value):
                        qty = parser.get_total_qty()
                else:
                    beg_ser = beg_ser_field.value or ''
                    end_ser = end_ser_field.value or ''
                    if beg_ser and end_ser:
                        parser = OORParser()
                        qty = parser.calculate_qty_from_range(beg_ser, end_ser)
                    else:
                        try:
                            qty = int(qty_field.value or 0)
                        except ValueError:
                            qty = 0
            except Exception:
                qty = 0
            
            try:
                unit_cost = float(unit_cost_field.value or 0.0)
            except ValueError:
                unit_cost = 0.0
            
            record_data = {
                'opco': opco,
                'device_type': device_type,
                'status': status_field.value or '',
                'mfr': mfr_field.value or '',
                'dev_code': dev_code_field. value or '',
                'beg_ser': beg_ser_field.value or '',
                'end_ser': end_ser_field.value or '',
                'oor_serial': oor_serial_value,
                'qty': qty,
                'po_date': po_date_field.value or '',
                'po_number': po_number_field.value or '',
                'recv_date': recv_date_field.value or '',
                'unit_cost': unit_cost,
                'cid': cid_field.value or '',
                'me_number': me_number_field.value or '',
                'pur_code': pur_code_field.value or '',
                'est': est_field.value or '',
                'use': use_field.value or '',
                'notes1': notes1_field. value or '',
                'notes2': notes2_field.value or '',
            }
            
            if is_edit:
                item_id = item. get('id')
                if self.db.update_item(item_id, record_data):
                    self. show_snackbar("Record updated successfully")
                else:
                    self.show_snackbar("Failed to update record", is_error=True)
            else:
                item_id = self.db.insert_item(record_data)
                if item_id > 0:
                    self.show_snackbar("Record added successfully")
                else:
                    self.show_snackbar("Failed to add record", is_error=True)
            
            dialog.open = False
            self.load_sheet_data(opco, device_type)
            self.update_dashboard_metrics()
            self.page.update()
        
        form_content = ft. Column(
            controls=[
                ft. Row([
                    ft.TextField(label="OpCo", value=opco, read_only=True, disabled=True, expand=True),
                    ft.TextField(label="Device Type", value=device_type, read_only=True, disabled=True, expand=True),
                ], spacing=16),
                ft. Row([status_field, mfr_field], spacing=16),
                ft. Row([dev_code_field, beg_ser_field, end_ser_field], spacing=16),
                oor_serial_field,
                ft.Text("Note: Qty will auto-calculate from OOR Serial or Beg/End Ser", size=11, italic=True, color=ft.Colors.ON_SURFACE_VARIANT),
                ft.Row([qty_field, po_date_field, po_number_field], spacing=16),
                ft.Row([recv_date_field, unit_cost_field, cid_field], spacing=16),
                ft. Row([me_number_field, pur_code_field, est_field, use_field], spacing=16),
                notes1_field,
                notes2_field,
            ],
            spacing=12,
            scroll=ft.ScrollMode.AUTO,
        )
        
        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text("Edit Record" if is_edit else "Add New Record"),
            content=ft.Container(content=form_content, width=700, height=500),
            actions=[
                ft.TextButton("Cancel", on_click=close_dialog),
                ft. FilledButton("Save", on_click=save_record),
            ],
            actions_alignment=ft.MainAxisAlignment. END,
        )
        
        self. page.overlay.append(dialog)
        dialog.open = True
        self.page.update()
    
    def delete_record(self, opco: str, device_type: str):
        """Delete selected record"""
        if not self.edit_mode:
            self. show_snackbar("Edit mode is locked. Enable edit mode first.", is_error=True)
            return
        
        sheet_name = f"{opco} - {device_type}"
        grid = self.sheet_grids.get(sheet_name)
        if not grid:
            return
        
        row_data = grid.get_selected_data()
        if not row_data:
            self.show_snackbar("Please select a record to delete", is_error=True)
            return
        
        item_id = row_data[0]
        
        def close_dialog(e):
            dialog.open = False
            self.page.update()
        
        def confirm_delete(e):
            if self.db.delete_item(item_id):
                self.show_snackbar("Record deleted successfully")
                self.load_sheet_data(opco, device_type)
                self.update_dashboard_metrics()
                self.selected_record_data = None
                self. update_detail_panel(None)
            else:
                self.show_snackbar("Failed to delete record", is_error=True)
            dialog.open = False
            self.page. update()
        
        dialog = ft. AlertDialog(
            modal=True,
            title=ft. Text("Confirm Delete"),
            content=ft.Text("Are you sure you want to delete this record? "),
            actions=[
                ft.TextButton("Cancel", on_click=close_dialog),
                ft.FilledButton("Delete", on_click=confirm_delete, style=ft.ButtonStyle(bgcolor=ft.Colors. ERROR)),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        self.page.overlay.append(dialog)
        dialog.open = True
        self.page.update()
    
    def import_data(self, opco: str, device_type: str):
        """Import data from file"""
        if not self.edit_mode:
            self.show_snackbar("Edit mode is locked. Enable edit mode first.", is_error=True)
            return
        
        def on_file_picked(e: ft.FilePickerResultEvent):
            if not e.files:
                return
            
            file_path = e. files[0].path
            if not file_path:
                return
            
            try:
                imported_count = 0
                
                if file_path.endswith('.xlsx'):
                    if not HAS_OPENPYXL:
                        self.show_snackbar("openpyxl is required for Excel import", is_error=True)
                        return
                    
                    wb = openpyxl. load_workbook(file_path)
                    sheet_name = f"{'OH' if opco == 'Ohio' else 'I&M'} - {device_type}"
                    
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if not any(row) or not row[0]:
                                continue
                            
                            if len(row) < 17:
                                continue
                            
                            unit_cost = 0.0
                            if len(row) > 11 and row[11]:
                                try:
                                    cost_val = row[11]
                                    if isinstance(cost_val, str):
                                        cost_str = cost_val. replace('$', '').replace(',', '')
                                        unit_cost = float(cost_str)
                                    else:
                                        unit_cost = float(cost_val)
                                except (ValueError, TypeError):
                                    unit_cost = 0.0
                            
                            qty = 0
                            if len(row) > 7 and row[7]:
                                try:
                                    qty = int(row[7])
                                except (ValueError, TypeError):
                                    qty = 0
                            
                            item = {
                                'opco': opco,
                                'device_type': device_type,
                                'status': str(row[2]) if len(row) > 2 and row[2] else '',
                                'mfr': str(row[3]) if len(row) > 3 and row[3] else '',
                                'dev_code': str(row[4]) if len(row) > 4 and row[4] else '',
                                'beg_ser': str(row[5]) if len(row) > 5 and row[5] else '',
                                'end_ser': str(row[6]) if len(row) > 6 and row[6] else '',
                                'oor_serial': '',
                                'qty': qty,
                                'po_date': str(row[8]) if len(row) > 8 and row[8] else '',
                                'po_number': str(row[9]) if len(row) > 9 and row[9] else '',
                                'recv_date': str(row[10]) if len(row) > 10 and row[10] else '',
                                'unit_cost': unit_cost,
                                'cid': str(row[12]) if len(row) > 12 and row[12] else '',
                                'me_number': str(row[13]) if len(row) > 13 and row[13] else '',
                                'pur_code': str(row[14]) if len(row) > 14 and row[14] else '',
                                'est': str(row[15]) if len(row) > 15 and row[15] else '',
                                'use': str(row[16]) if len(row) > 16 and row[16] else '',
                                'notes1': str(row[17]) if len(row) > 17 and row[17] else '',
                                'notes2': str(row[18]) if len(row) > 18 and row[18] else '',
                            }
                            
                            result = self.db. insert_item(item)
                            if result > 0:
                                imported_count += 1
                
                self.load_sheet_data(opco, device_type)
                self.update_dashboard_metrics()
                self.show_snackbar(f"Imported {imported_count} records")
                
            except Exception as ex:
                logger.error(f"Error importing file: {ex}")
                self.show_snackbar(f"Error importing file: {ex}", is_error=True)
        
        file_picker = ft.FilePicker(on_result=on_file_picked)
        self.page.overlay. append(file_picker)
        self.page.update()
        file_picker.pick_files(
            allowed_extensions=["xlsx", "csv"],
            dialog_title="Select File to Import",
        )
    
    def export_data(self, opco: str, device_type: str):
        """Export data to CSV"""
        def on_file_picked(e: ft. FilePickerResultEvent):
            if not e.path:
                return
            
            file_path = e.path
            if not file_path. endswith('.csv'):
                file_path += '.csv'
            
            try:
                items = self.db. get_items_by_sheet(opco, device_type)
                
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(self.FULL_COLUMNS)
                    writer.writerows(items)
                
                self.show_snackbar(f"Exported {len(items)} records to {file_path}")
                
            except Exception as ex:
                self.show_snackbar(f"Error exporting: {ex}", is_error=True)
        
        file_picker = ft.FilePicker(on_result=on_file_picked)
        self.page.overlay.append(file_picker)
        self.page.update()
        file_picker.save_file(
            file_name=f"{opco}_{device_type}. csv",
            dialog_title="Export CSV",
        )
    
    def show_statistics(self, opco: str, device_type: str):
        """Show statistics dialog"""
        try:
            stats = self.db.get_statistics(opco, device_type)
            
            stats_content = ft.Column(
                controls=[
                    ft. Row([ft.Text("Total Records:", weight=ft.FontWeight.W_500), ft.Container(expand=True), ft.Text(f"{stats['total_items']:,}")]),
                    ft.Row([ft.Text("Total Quantity:", weight=ft. FontWeight.W_500), ft.Container(expand=True), ft.Text(f"{stats['total_qty']:,}")]),
                    ft. Row([ft.Text("Total Value:", weight=ft. FontWeight.W_500), ft.Container(expand=True), ft.Text(f"${stats['total_value']:,.2f}")]),
                    ft.Row([ft. Text("Average Cost per Unit:", weight=ft.FontWeight.W_500), ft. Container(expand=True), ft.Text(f"${stats['avg_cost']:,.2f}")]),
                ],
                spacing=12,
            )
            
            def close_dialog(e):
                dialog.open = False
                self.page.update()
            
            dialog = ft.AlertDialog(
                title=ft.Text(f"Statistics - {opco} {device_type}"),
                content=ft.Container(content=stats_content, width=350),
                actions=[ft.TextButton("Close", on_click=close_dialog)],
            )
            
            self.page.overlay.append(dialog)
            dialog.open = True
            self.page.update()
            
        except Exception as e:
            self.show_snackbar(f"Error getting statistics: {e}", is_error=True)
    
    def toggle_theme(self, e):
        """Toggle between light and dark theme"""
        if self.page.theme_mode == ft.ThemeMode.LIGHT:
            self. page.theme_mode = ft.ThemeMode.DARK
            self.current_theme = "Dark"
            if hasattr(e. control, 'icon'):
                e. control.icon = ft.Icons.DARK_MODE
        else:
            self.page.theme_mode = ft.ThemeMode.LIGHT
            self.current_theme = "Light"
            if hasattr(e.control, 'icon'):
                e.control.icon = ft.Icons. LIGHT_MODE
        self.page.update()
    
    def show_about(self, e):
        """Show about dialog"""
        about_content = ft.Column(
            controls=[
                ft. Text("Created Histories v2. 0. 0 (Flet)", weight=ft.FontWeight. BOLD, size=16),
                ft. Container(height=12),
                ft.Text("Modern Professional Device History Tracking System", size=14),
                ft.Container(height=12),
                ft.Text("Features:", weight=ft.FontWeight.W_600),
                ft.Text("• Modern dashboard with real-time metrics"),
                ft. Text("• Multi-sheet support with search"),
                ft.Text("• CRUD operations with password protection"),
                ft.Text("• Import/Export functionality"),
                ft.Text("• Light/Dark theme support"),
                ft. Container(height=12),
                ft.Text("Built with Flet (Flutter for Python)", size=12, color=ft.Colors.ON_SURFACE_VARIANT),
            ],
            spacing=2,
        )
        
        def close_dialog(e):
            dialog.open = False
            self.page.update()
        
        dialog = ft.AlertDialog(
            title=ft. Text("About"),
            content=ft.Container(content=about_content, width=400),
            actions=[ft.TextButton("Close", on_click=close_dialog)],
        )
        
        self.page.overlay.append(dialog)
        dialog.open = True
        self.page. update()
    
    def show_help(self, e):
        """Show help dialog"""
        help_content = ft. Column(
            controls=[
                ft. Text("How to Use:", weight=ft.FontWeight. W_600, size=16),
                ft.Container(height=8),
                ft. Text("1. Navigate using the rail on the left"),
                ft.Text("2. Click 'Enable Edit Mode' and enter password to edit"),
                ft.Text("3. Use toolbar buttons to add, edit, or delete records"),
                ft.Text("4.  Use the search bar to filter data"),
                ft. Text("5. Import Excel files or export to CSV"),
                ft.Container(height=12),
                ft. Text("Default Password: admin123", weight=ft.FontWeight.W_500),
            ],
            spacing=4,
        )
        
        def close_dialog(e):
            dialog.open = False
            self.page.update()
        
        dialog = ft.AlertDialog(
            title=ft.Text("Help"),
            content=ft.Container(content=help_content, width=400),
            actions=[ft. TextButton("Close", on_click=close_dialog)],
        )
        
        self.page. overlay.append(dialog)
        dialog. open = True
        self.page.update()
    
    def show_snackbar(self, message: str, is_error: bool = False):
        """Show a snackbar notification"""
        snackbar = ft. SnackBar(
            content=ft. Text(message),
            bgcolor=ft.Colors. ERROR if is_error else None,
            action="OK",
        )
        self.page.overlay.append(snackbar)
        snackbar.open = True
        self. page.update()


def main(page: ft.Page):
    """Main application entry point"""
    logger.info("=" * 60)
    logger.info("Created Histories Module (Flet)")
    logger.info(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=" * 60)
    
    CreatedHistoriesApp(page)


if __name__ == '__main__':
    ft.app(target=main)